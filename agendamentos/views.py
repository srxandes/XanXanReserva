# ─── Imports ──────────────────────────────────────────────────────────────────
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.shortcuts import redirect
from django.core.exceptions import ValidationError
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
from .forms import AgendamentoForm
from .models import Configuracao, Agendamento, EmailBloqueado, LogAtividade


# ─── Página inicial ───────────────────────────────────────────────────────────
def index(request):
    return render(request, 'agendamentos/index.html')


# ─── Formulário de agendamento ────────────────────────────────────────────────
def fazer_agendamento(request):
    # Tarefas automáticas executadas a cada acesso ao formulário
    Agendamento.expirar_pendentes()       # Rejeita pendentes vencidos
    Agendamento.enviar_lembretes_do_dia() # Envia lembretes de amanhã

    config = Configuracao.objects.first()
    instrucoes = config.texto_instrucoes if config else "Preencha as informações abaixo."

    # Mensagens de erro falsas para e-mails bloqueados (não revelam o bloqueio)
    MENSAGENS_ERRO = [
        "Não foi possível continuar, erro 500.",
        "Falha de conexão com o servidor. Tente novamente mais tarde.",
        "Tempo limite de requisição esgotado. Código: TIMEOUT_408.",
        "Serviço temporariamente indisponível. Erro 503.",
        "Ocorreu um erro inesperado. Por favor, tente mais tarde.",
        "Falha interna no processamento. Código: ERR_INTERNAL.",
        "Não foi possível completar a operação. Erro 502.",
    ]

    if request.method == 'POST':
        form = AgendamentoForm(request.POST)

        # Bloqueia e-mails na lista negra com mensagem genérica aleatória
        email_digitado = request.POST.get('email', '').strip().lower()
        if EmailBloqueado.objects.filter(email__iexact=email_digitado).exists():
            messages.error(request, random.choice(MENSAGENS_ERRO))
            form = AgendamentoForm()
            aprovados  = Agendamento.objects.filter(status='aprovado', entregue=False)
            pendentes  = Agendamento.objects.filter(status='pendente', entregue=False, equipamento='notebook')
            antecedencia = config.antecedencia_minima if config else 30
            return render(request, 'agendamentos/formulario.html', {
                'form': form,
                'instrucoes': instrucoes,
                'agendamentos_aprovados': aprovados,
                'agendamentos_pendentes': pendentes,
                'antecedencia_minima': antecedencia,
            })

        if form.is_valid():
            try:
                agendamento = form.save()
                # refresh_from_db garante que o status definido no save() do model seja lido corretamente
                agendamento.refresh_from_db()
                return render(request, 'agendamentos/sucesso.html', {
                    'agendamento': agendamento,
                })
            except ValidationError as e:
                # Erros de validação do model (ex: conflito de estoque, antecedência)
                if hasattr(e, 'message_dict'):
                    for field, errors in e.message_dict.items():
                        form.add_error(field, errors)
                else:
                    form.add_error(None, e.message)
            except Exception as e:
                messages.error(request, f"Ocorreu um erro técnico: {str(e)}. Por favor, informe ao NTI.")
        else:
            messages.warning(request, "Existem erros no preenchimento. Verifique os campos em vermelho.")
    else:
        form = AgendamentoForm()

    # Passa agendamentos existentes ao template para o JS bloquear horários cheios
    aprovados    = Agendamento.objects.filter(status='aprovado', entregue=False)
    pendentes    = Agendamento.objects.filter(status='pendente', entregue=False, equipamento='notebook')
    antecedencia = config.antecedencia_minima if config else 30

    return render(request, 'agendamentos/formulario.html', {
        'form': form,
        'instrucoes': instrucoes,
        'agendamentos_aprovados': aprovados,
        'agendamentos_pendentes': pendentes,
        'antecedencia_minima': antecedencia,
    })


# ─── Agenda pública ───────────────────────────────────────────────────────────
def ver_agenda(request):
    return render(request, 'agendamentos/agenda.html')


# ─── Helpers de permissão ─────────────────────────────────────────────────────
def is_superuser(user):
    """Permite acesso apenas a superusuários."""
    return user.is_superuser

def is_tecno_camb(user):
    """Permite acesso apenas ao usuário tecno_camb (dono do sistema)."""
    return user.is_superuser and user.username == 'tecno_camb'


# ─── Cadastrar admin ──────────────────────────────────────────────────────────
@login_required
@user_passes_test(is_superuser)
def cadastrar_admin(request):
    erro   = None
    sucesso = None

    if request.method == 'POST':
        username  = request.POST.get('username', '').strip()
        email     = request.POST.get('email', '').strip()
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')

        # Validações antes de criar
        if not username or not password1:
            erro = "Usuário e senha são obrigatórios."
        elif password1 != password2:
            erro = "As senhas não coincidem."
        elif len(password1) < 8:
            erro = "A senha deve ter pelo menos 8 caracteres."
        elif User.objects.filter(username=username).exists():
            erro = f"O usuário '{username}' já existe."
        else:
            User.objects.create_superuser(username=username, email=email, password=password1)
            sucesso = f"Admin '{username}' criado com sucesso!"
            LogAtividade.registrar('admin_criado', f"Superusuário '{username}' criado.")

    # Lista de admins exibida apenas para tecno_camb (para remoção)
    admins       = User.objects.filter(is_superuser=True).exclude(username=request.user.username)
    pode_deletar = request.user.username == 'tecno_camb'

    return render(request, 'agendamentos/cadastro_admin.html', {
        'erro': erro,
        'sucesso': sucesso,
        'admins': admins,
        'pode_deletar': pode_deletar,
    })


# ─── Deletar admin ────────────────────────────────────────────────────────────
@login_required
@user_passes_test(is_tecno_camb)  # Apenas tecno_camb pode deletar
def deletar_admin(request, user_id):
    if request.method == 'POST':
        try:
            usuario = User.objects.get(id=user_id, is_superuser=True)
            if usuario.username == 'tecno_camb':
                # Proteção: tecno_camb não pode se auto-deletar
                messages.error(request, "Não é possível deletar o usuário tecno_camb.")
            else:
                nome = usuario.username
                usuario.delete()
                LogAtividade.registrar('admin_criado', f"Superusuário '{nome}' removido por {request.user.username}.")
                messages.success(request, f"Admin '{nome}' removido com sucesso.")
        except User.DoesNotExist:
            messages.error(request, "Usuário não encontrado.")
    return redirect('cadastrar_admin')


# ─── Exportar histórico para Excel ───────────────────────────────────────────
@login_required
def exportar_historico_excel(request):
    """Gera e retorna um arquivo .xlsx com o histórico de agendamentos.
    Respeita os mesmos filtros do dashboard (período, equipamento, status)."""
    filtro_inicio = request.GET.get('data_inicio', '')
    filtro_fim    = request.GET.get('data_fim', '')
    filtro_equip  = request.GET.get('equipamento', '')
    filtro_status = request.GET.get('status', '')

    # Aplica os filtros recebidos via GET
    qs = Agendamento.objects.all().order_by('-data_inicio')
    if filtro_inicio:
        qs = qs.filter(data_inicio__date__gte=filtro_inicio)
    if filtro_fim:
        qs = qs.filter(data_inicio__date__lte=filtro_fim)
    if filtro_equip:
        qs = qs.filter(equipamento=filtro_equip)
    if filtro_status:
        qs = qs.filter(status=filtro_status)

    # ── Monta o arquivo Excel ─────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Histórico"

    # Estilos
    header_fill  = PatternFill("solid", start_color="2C4A5E")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font    = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    border_side  = Side(style="thin", color="D1D5DB")
    cell_border  = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # Cores de fundo por status
    status_cores = {
        "aprovado":  "D1FAE5",  # Verde claro
        "pendente":  "FEF3C7",  # Amarelo claro
        "rejeitado": "FEE2E2",  # Vermelho claro
    }
    status_labels = {
        "aprovado":  "Aprovado",
        "pendente":  "Pendente",
        "rejeitado": "Rejeitado",
    }

    # Cabeçalho
    colunas = ["#", "Nome", "E-mail", "Turma", "Equipamento", "Quantidade", "Início", "Fim", "Status", "Entregue"]
    for col, titulo in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = cell_border
    ws.row_dimensions[1].height = 30

    # Linhas de dados
    for i, ag in enumerate(qs, start=1):
        row = i + 1
        valores = [
            i,
            ag.nome,
            ag.email,
            ag.turma,
            ag.get_equipamento_display(),
            ag.quantidade,
            ag.data_inicio.strftime("%d/%m/%Y %H:%M") if ag.data_inicio else "",
            ag.data_fim.strftime("%d/%m/%Y %H:%M")    if ag.data_fim    else "",
            status_labels.get(ag.status, ag.status),
            "Sim" if ag.entregue else "Não",
        ]
        for col, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.font      = cell_font
            cell.border    = cell_border
            cell.alignment = center_align if col != 2 else Alignment(vertical="center")
            # Coluna 9 = Status: aplica cor de fundo
            if col == 9:
                cor = status_cores.get(ag.status, "FFFFFF")
                cell.fill = PatternFill("solid", start_color=cor)
        ws.row_dimensions[row].height = 18

    # Linha de total (fórmula Excel — recalcula automaticamente)
    total_row = len(list(qs)) + 2
    ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True, name="Arial", size=10)
    ws.cell(row=total_row, column=6, value=f'=SUM(F2:F{total_row - 1})').font = Font(bold=True, name="Arial", size=10)

    # Larguras das colunas
    larguras = [5, 28, 28, 12, 14, 10, 18, 18, 14, 10]
    for col, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col)].width = largura

    # Salva em memória e retorna como download
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="historico_agendamentos.xlsx"'
    return response